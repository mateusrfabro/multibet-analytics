"""
Report: RTP observado de TODOS os jogos desde o D0 da operacao Play4Tune.

Demanda Head (20/04): "quero ver como ta o RTP de todos os jogos, 100%,
nos ultimos 30d. desde o D0 da operacao."

D0 da operacao = 27/03/2026 (primeira aposta em casino_user_game_metrics).
DATA_FIM = ontem (D-1). Regra: entregas a lideranca sao SEMPRE D-1 (memory/feedback_sempre_usar_d_menos_1.md).

Fontes:
  - casino_user_game_metrics  → rodadas, turnover, win (calcula rtp_obs)
  - casino_games.rtp          → RTP do catalogo (pode estar defasado vs 2J)
  - RTP_OFICIAL_CSV_PATH      → planilha oficial 2J Games (source of truth)
    Ver memory/project_rtp_oficial_2j_sheet.md

Filtro: test users UNION + whitelist DP/SQ

Saidas:
  - stdout: ranking por volume com RTP obs e veredito
  - reports/play4_rtp_todos_jogos_desde_d0_FINAL.xlsx (3 abas)
"""
import os, sys
from datetime import date, datetime, timedelta

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from db.supernova_bet import get_supernova_bet_connection
import report_players_afundando_ggr_play4 as mainmod
from report_players_afundando_ggr_play4 import (
    get_test_user_ids, fetch_fx_rates, to_brl, fmt_int,
)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


D0_OPERACAO = date(2026, 3, 27)
DATA_FIM    = date.today() - timedelta(days=1)  # D-1 obrigatorio p/ lideranca

OUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "reports")
XLSX_PATH    = os.path.join(OUT_DIR, "play4_rtp_todos_jogos_desde_d0_FINAL.xlsx")
CSV_PATH     = os.path.join(OUT_DIR, "play4_rtp_todos_jogos_desde_d0_FINAL.csv")
LEGENDA_PATH = os.path.join(OUT_DIR, "play4_rtp_todos_jogos_desde_d0_FINAL_legenda.txt")

# Planilha oficial 2J Games (source of truth p/ RTP config).
# Exportar Google Sheets como CSV e salvar neste caminho (pode ser atualizado).
# Link: https://docs.google.com/spreadsheets/d/1jU7XNjp02nmp0A29tnCUbw01gHsN9AIM/edit
RTP_OFICIAL_CSV_PATH = os.path.join(OUT_DIR, "rtp_oficial_2j_games.csv")


def load_rtp_oficial_2j():
    """Le CSV com RTP oficial do 2J Games.
    Retorna (by_external_id, by_name, fonte) — ambos dicts apontando pro mesmo rtp.
    Match no script usa external_id (preferencial, deterministico) ou nome (fallback).
    Se arquivo nao existir, retorna ({}, {}, None) e o script usa rtp_cfg do banco.
    """
    if not os.path.exists(RTP_OFICIAL_CSV_PATH):
        return {}, {}, None
    import csv
    by_eid = {}; by_name = {}
    with open(RTP_OFICIAL_CSV_PATH, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            eid = None; nome = None; rtp = None
            for k, v in row.items():
                kl = (k or "").strip().lower()
                if kl in ("external_id", "game_id", "id", "provider_id_game"):
                    eid = (str(v) if v is not None else "").strip()
                elif kl in ("name", "nome", "jogo", "game", "game_name"):
                    nome = (v or "").strip().upper()
                elif kl in ("rtp", "rtp_oficial", "rtp_config", "rtp_target", "rtp_pct"):
                    try:
                        rtp = float(str(v).replace(",", ".").replace("%", "").strip())
                    except (ValueError, TypeError):
                        rtp = None
            if rtp is None:
                continue
            if eid:
                by_eid[eid] = rtp
            if nome:
                by_name[nome] = rtp
    return by_eid, by_name, RTP_OFICIAL_CSV_PATH


SQL_RTP_TODOS = """
SELECT
    g.name AS jogo,
    g.external_id AS external_id,
    COALESCE(pv.name, 'Unknown') AS provider,
    g.rtp AS rtp_cfg,
    g.active AS jogo_ativo,
    COUNT(DISTINCT m.date) AS dias_ativo,
    COUNT(DISTINCT m.user_id) AS players_distintos,
    SUM(m.played_rounds) AS rodadas,
    ROUND(SUM(m.total_bet_amount)::numeric, 2) AS apostado,
    ROUND(SUM(m.total_win_amount)::numeric, 2) AS ganho,
    ROUND(SUM(m.net_revenue)::numeric, 2) AS ggr,
    ROUND(CASE WHEN SUM(m.total_bet_amount) > 0
               THEN SUM(m.total_win_amount) / SUM(m.total_bet_amount) * 100
               ELSE NULL END::numeric, 1) AS rtp_obs
FROM casino_user_game_metrics m
JOIN casino_games g ON g.id = m.game_id
LEFT JOIN casino_providers pv ON pv.id = g.provider_id
WHERE m.date BETWEEN %s AND %s
  AND m.user_id NOT IN %s
GROUP BY g.name, g.external_id, pv.name, g.rtp, g.active
ORDER BY SUM(m.played_rounds) DESC
"""


def classificar(rtp_obs, rtp_cfg, rodadas):
    if rtp_obs is None:
        return ("sem dado", None)
    if rodadas is None or rodadas < 500:
        return ("amostra baixa (<500 giros)", None)
    delta = float(rtp_obs) - float(rtp_cfg or 0)
    if delta > 15:
        return ("ANOMALO (+15pp)", delta)
    if delta > 5:
        return ("alto (+5pp)", delta)
    if delta < -5:
        return ("abaixo (casa ganha)", delta)
    return ("normal", delta)


def fmt_rs(v):
    if v is None: return "—"
    return f"Rs {float(v):>12,.0f}"


def fmt_brl_rs(v):
    if v is None: return "—"
    return f"R$ {to_brl(v):,.2f} ({fmt_rs(v).strip()})"


def run():
    total_dias = (DATA_FIM - D0_OPERACAO).days + 1
    print("=" * 90)
    print("PLAY4TUNE — RTP observado TODOS os jogos (desde D0 da operacao)")
    print(f"Janela: {D0_OPERACAO.strftime('%d/%m/%Y')} a {DATA_FIM.strftime('%d/%m/%Y')} "
          f"({total_dias} dias)")
    print(f"Rodado em: {datetime.now().strftime('%d/%m/%Y %H:%M')} BRT")
    print("=" * 90)

    tunnel, conn = get_supernova_bet_connection()
    conn.set_session(readonly=True, autocommit=True)
    cur = conn.cursor()
    try:
        pkr_to_brl, fx_at = fetch_fx_rates(cur)
        mainmod.PKR_TO_BRL = pkr_to_brl
        mainmod.FX_FETCHED_AT = fx_at
        print(f"[FX] 1 PKR = R$ {pkr_to_brl:.6f} (snapshot {fx_at} UTC)")

        test_ids, n_test, _ = get_test_user_ids(cur)
        print(f"[Filtro] {n_test} contas teste excluidas")

        rtp_by_eid, rtp_by_name, rtp_oficial_fonte = load_rtp_oficial_2j()
        if rtp_oficial_fonte:
            print(f"[RTP 2J] {len(rtp_by_eid)} entries por external_id, "
                  f"{len(rtp_by_name)} por nome — fonte: "
                  f"{os.path.basename(rtp_oficial_fonte)}\n")
        else:
            print(f"[RTP 2J] !! planilha oficial NAO encontrada em {RTP_OFICIAL_CSV_PATH}\n"
                  f"          script seguira usando apenas rtp do banco (casino_games.rtp)\n")

        cur.execute(SQL_RTP_TODOS, (D0_OPERACAO, DATA_FIM, test_ids))
        rows = cur.fetchall()

        registros = []
        for r in rows:
            jogo, external_id, provider, rtp_cfg, ativo, dias_ativo, players, rodadas, \
                apostado, ganho, ggr, rtp_obs = r
            rtp_cfg_f = float(rtp_cfg or 0)
            rtp_obs_f = float(rtp_obs) if rtp_obs is not None else None
            # Match 2J: prioriza external_id (deterministico), fallback = nome
            rtp_oficial = rtp_by_eid.get(str(external_id) if external_id else "")
            if rtp_oficial is None:
                rtp_oficial = rtp_by_name.get((jogo or "").upper())
            # RTP de referencia: prefere planilha oficial 2J; fallback = banco
            rtp_ref = rtp_oficial if rtp_oficial is not None else rtp_cfg_f
            veredito, delta = classificar(rtp_obs_f, rtp_ref, rodadas or 0)
            delta_banco_vs_oficial = (
                rtp_cfg_f - rtp_oficial if rtp_oficial is not None else None
            )
            registros.append({
                "jogo": jogo, "external_id": external_id, "provider": provider,
                "rtp_cfg": rtp_cfg_f, "ativo": ativo,
                "rtp_oficial": rtp_oficial,
                "delta_banco_vs_oficial": delta_banco_vs_oficial,
                "dias_ativo": dias_ativo, "players": players,
                "rodadas": rodadas or 0,
                "apostado": float(apostado or 0),
                "ganho": float(ganho or 0),
                "ggr": float(ggr or 0),
                "rtp_obs": float(rtp_obs) if rtp_obs is not None else None,
                "delta": delta, "veredito": veredito,
            })

        # Sumarios
        total_jogos = len(registros)
        com_amostra = [r for r in registros if r["rodadas"] >= 500]
        anomalos    = [r for r in com_amostra if r["veredito"].startswith("ANOMALO")]
        altos       = [r for r in com_amostra if r["veredito"].startswith("alto")]
        abaixo      = [r for r in com_amostra if r["veredito"].startswith("abaixo")]
        normais     = [r for r in com_amostra if r["veredito"] == "normal"]
        amostra_baixa = [r for r in registros if r["rodadas"] < 500]

        # GGR total operacao
        ggr_total = sum(r["ggr"] for r in registros)
        turnover_total = sum(r["apostado"] for r in registros)

        print(f"Resumo da operacao no periodo:")
        print(f"  Jogos com apostas: {total_jogos} ({len(com_amostra)} c/ amostra >=500 giros)")
        print(f"  Turnover total: {fmt_brl_rs(turnover_total)}")
        print(f"  GGR casa total: {fmt_brl_rs(ggr_total)}")
        print(f"  Veredito: {len(anomalos)} ANOMALOS, {len(altos)} altos, "
              f"{len(normais)} normais, {len(abaixo)} abaixo, "
              f"{len(amostra_baixa)} amostra baixa")

        # Print ordenado por delta DESC (pior RTP primeiro)
        def print_bloco(titulo, lista, ordenar_por_delta_desc=True):
            if not lista: return
            print(f"\n{'=' * 90}")
            print(f"{titulo}")
            print('=' * 90)
            if ordenar_por_delta_desc:
                lista = sorted(lista, key=lambda x: -(x["delta"] or -999))
            print(f"{'Jogo':<28} {'RTP cfg':>8} {'RTP obs':>8} {'Delta':>7} "
                  f"{'Rodadas':>8} {'Players':>7} {'Apostado (Rs)':>14} {'GGR (Rs)':>12}")
            print("-" * 100)
            for r in lista:
                delta_str = f"{r['delta']:+.1f}pp" if r['delta'] is not None else "—"
                rtp_str = f"{r['rtp_obs']:.1f}%" if r['rtp_obs'] is not None else "—"
                print(f"{r['jogo'][:28]:<28} {r['rtp_cfg']:>7.1f}% {rtp_str:>8} "
                      f"{delta_str:>7} {r['rodadas']:>8,} {r['players']:>7} "
                      f"{r['apostado']:>14,.0f} {r['ggr']:>12,.0f}")

        print_bloco("ANOMALOS (delta > +15pp) — revisar com provider", anomalos)
        print_bloco("ALTOS (delta +5 a +15pp) — monitorar", altos)
        print_bloco("ABAIXO do config (casa ganha mais que o esperado)", abaixo)
        print_bloco("NORMAIS (oscilacao natural)", normais[:10])
        if len(normais) > 10:
            print(f"  ... + {len(normais) - 10} normais (ver Excel)")

        # Desalinhamento banco vs planilha oficial 2J — catalogo defasado?
        if rtp_by_eid or rtp_by_name:
            desalinhados = [r for r in registros
                            if r.get("delta_banco_vs_oficial") is not None
                            and abs(r["delta_banco_vs_oficial"]) > 0.01]
            sem_oficial = [r for r in registros if r.get("rtp_oficial") is None]
            print(f"\n{'=' * 90}")
            print(f"SYNC BANCO vs PLANILHA OFICIAL 2J")
            print('=' * 90)
            print(f"  Jogos com RTP oficial: {len(registros) - len(sem_oficial)}/{len(registros)}")
            print(f"  Jogos desalinhados (banco != oficial): {len(desalinhados)}")
            if sem_oficial:
                print(f"  Jogos SEM RTP na planilha (verificar cobertura da planilha):")
                for r in sem_oficial[:10]:
                    print(f"    - {r['jogo']}")
                if len(sem_oficial) > 10:
                    print(f"    ... + {len(sem_oficial) - 10} (ver Excel coluna 'RTP oficial 2J')")
            if desalinhados:
                print(f"  Top desalinhamentos (|delta| desc) — catalogo do banco precisa sync:")
                for r in sorted(desalinhados,
                                key=lambda x: -abs(x["delta_banco_vs_oficial"]))[:15]:
                    print(f"    {r['jogo'][:30]:<30}  "
                          f"banco={r['rtp_cfg']:>6.2f}%  oficial={r['rtp_oficial']:>6.2f}%  "
                          f"delta={r['delta_banco_vs_oficial']:+.2f}pp")

        if HAS_OPENPYXL:
            gerar_excel(registros, ggr_total, turnover_total, total_dias)
            print(f"\n[Excel] Salvo em: {XLSX_PATH}")

        gerar_csv(registros)
        print(f"[CSV]      Salvo em: {CSV_PATH}")
        print(f"[Legenda]  Salvo em: {LEGENDA_PATH}")

    finally:
        cur.close(); conn.close(); tunnel.stop()


def gerar_csv(registros):
    """CSV com todos os 122 jogos, ordenado por delta desc dentro de cada veredito."""
    import csv

    ordenados = sorted(registros, key=lambda r: (
        0 if r["veredito"].startswith("ANOMALO") else
        1 if r["veredito"].startswith("alto") else
        2 if r["veredito"].startswith("abaixo") else
        3 if r["veredito"] == "normal" else 4,
        -(r["delta"] or -999),
    ))

    with open(CSV_PATH, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow([
            "rank", "jogo", "provider", "ativo",
            "rtp_oficial_2j_pct", "rtp_cfg_banco_pct", "delta_banco_vs_oficial_pp",
            "rtp_obs_pct", "delta_pp",
            "dias_ativo", "players_distintos", "rodadas",
            "apostado_brl", "apostado_pkr",
            "ganho_brl",    "ganho_pkr",
            "ggr_brl",      "ggr_pkr",
            "veredito",
        ])
        for i, r in enumerate(ordenados, start=1):
            w.writerow([
                i, r["jogo"], r["provider"], "Sim" if r["ativo"] else "Nao",
                f"{r['rtp_oficial']:.2f}" if r.get("rtp_oficial") is not None else "",
                f"{r['rtp_cfg']:.2f}",
                f"{r['delta_banco_vs_oficial']:+.2f}" if r.get("delta_banco_vs_oficial") is not None else "",
                f"{r['rtp_obs']:.2f}" if r["rtp_obs"] is not None else "",
                f"{r['delta']:+.2f}" if r["delta"] is not None else "",
                r["dias_ativo"], r["players"], r["rodadas"],
                f"{to_brl(r['apostado']):.2f}", f"{r['apostado']:.2f}",
                f"{to_brl(r['ganho']):.2f}",    f"{r['ganho']:.2f}",
                f"{to_brl(r['ggr']):.2f}",      f"{r['ggr']:.2f}",
                r["veredito"],
            ])

    with open(LEGENDA_PATH, "w", encoding="utf-8") as f:
        f.write(f"""LEGENDA — play4_rtp_todos_jogos_desde_d0_FINAL.csv
{'=' * 70}

CONTEXTO
--------
Demanda: Head (20/04/2026) — "quero ver como ta o RTP de todos os jogos,
100%, nos ultimos 30d. desde o D0 da operacao."

Janela: {D0_OPERACAO.strftime('%d/%m/%Y')} a {DATA_FIM.strftime('%d/%m/%Y')}
        ({(DATA_FIM - D0_OPERACAO).days + 1} dias — operacao tem 25d, "ultimos 30d" == "desde D0")

Fontes:
  - casino_user_game_metrics (supernova_bet, validada 09/04/2026) — rodadas, win, bet
  - casino_games.rtp          — RTP do catalogo do banco (pode estar defasado)
  - rtp_oficial_2j_games.csv  — planilha oficial 2J Games, source of truth
    (sheet: https://docs.google.com/spreadsheets/d/1jU7XNjp02nmp0A29tnCUbw01gHsN9AIM)
Filtro: 72 contas teste excluidas (heuristica + logica dev), 4 whitelist DP/SQ

Separador CSV: ponto-e-virgula (;)
Encoding: UTF-8 com BOM (abre no Excel BR sem acentos quebrados)
Casas decimais: ponto (.)

COLUNAS
-------
rank               Ordem: ANOMALOS primeiro, depois altos, abaixo, normais, amostra baixa.
                   Dentro de cada grupo, maior delta primeiro.
jogo               Nome do jogo conforme catalogo do provider.
provider           Provider ativo — atualmente 100% "2J Game".
ativo              Sim/Nao — se o jogo esta marcado como ativo no catalogo.
rtp_oficial_2j_pct        RTP oficial informado pela 2J Games na planilha (source of truth).
                          Vazio se o jogo nao consta na planilha do provider.
rtp_cfg_banco_pct         RTP gravado em supernova_bet.casino_games.rtp
                          (snapshot do banco — pode estar defasado vs 2J).
delta_banco_vs_oficial_pp rtp_cfg_banco - rtp_oficial_2j. Positivo/negativo alto =
                          catalogo do banco desatualizado, precisa sync com provider.
rtp_obs_pct               RTP observado = ganho / apostado x 100 (em %). Vazio se sem apostas.
delta_pp                  RTP observado - RTP de referencia (oficial 2J se disponivel;
                          senao, rtp do banco).
                          Positivo = casa devolveu mais que o esperado.
                          Negativo = casa ganhou mais que o esperado.
dias_ativo         Dias distintos com pelo menos 1 aposta no periodo.
players_distintos  Jogadores reais distintos que apostaram.
rodadas            Total de giros/spins no periodo.
apostado_brl       Turnover em BRL (conversao triangulada via USD).
apostado_pkr       Turnover na moeda original da operacao (Rupia Paquistanesa).
ganho_brl          Valor pago aos jogadores em BRL.
ganho_pkr          Valor pago em PKR.
ggr_brl            Gross Gaming Revenue da casa em BRL (apostado - ganho). Negativo = casa perdeu.
ggr_pkr            GGR em PKR.
veredito           Classificacao — ver criterios abaixo.

CRITERIOS DE VEREDITO
---------------------
ANOMALO (+15pp)        Delta > +15pp com >=500 giros. Revisar urgente com provider.
alto (+5pp)            Delta entre +5 e +15pp. Monitorar — pode ser tendencia ou volatilidade.
acima / abaixo         Delta moderado. Operacao normal.
abaixo (casa ganha)    Delta < -5pp. Casa esta ganhando mais que o config prometia —
                       pode indicar RTP cfg errado pra cima (revisar) ou jogo com
                       caracteristica de baixa variancia.
normal                 Delta entre -5 e +5pp. Oscilacao natural dentro do esperado.
amostra baixa          < 500 giros no periodo inteiro. Ignorar — ruido estatistico alto.

POR QUE >=500 GIROS?
Com RTP tipico ~95% e variancia alta por jogada, 500 giros dao desvio padrao
de ~5pp no RTP observado. Abaixo disso o ruido mata qualquer conclusao — mesmo
um RTP config 95% pode "observar" 120% ou 70% por puro azar/sorte do jogador.

MOEDA E CONVERSAO
-----------------
Operacao Play4Tune roda 100% em PKR (Rupia Paquistanesa).
BRL (Real) e convertido via triangulacao USD usando a tabela
`currency_exchange_rates` do proprio supernova_bet (fonte: openexchangerate-free,
atualizada 2x/dia).

Taxa usada neste relatorio: 1 PKR = R$ {mainmod.PKR_TO_BRL:.6f}
Snapshot: {mainmod.FX_FETCHED_AT.strftime('%d/%m/%Y %H:%M UTC') if mainmod.FX_FETCHED_AT else '—'}

COMO INTERPRETAR
----------------
1. Jogos com delta DENTRO de -5 a +5pp e amostra >=500 giros = operacao saudavel.
2. Jogos ANOMALOS ou ALTOS = investigar:
   - 1 jogador apostando muito pontualmente (volatilidade) ou
   - bug/config errado do provider.
3. Jogos com amostra baixa = insuficiente pra concluir. Revisitar quando
   acumular volume.
4. Para diagnostico PRODUTO (jogo quebrado?): janela >=30d.
   Para DETECCAO DE EVENTO (jogador suspeito agora?): janela <=7d.

GERADO EM: {datetime.now().strftime('%d/%m/%Y %H:%M')} BRT
SCRIPT: scripts/report_rtp_todos_jogos_desde_d0.py
""")


def gerar_excel(registros, ggr_total, turnover_total, total_dias):
    wb = openpyxl.Workbook()
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    anomalo_fill = PatternFill("solid", fgColor="FEE2E2")
    alto_fill    = PatternFill("solid", fgColor="FEF3C7")
    abaixo_fill  = PatternFill("solid", fgColor="D1FAE5")

    # Aba 1: Resumo executivo
    ws = wb.active
    ws.title = "Resumo"
    ws["A1"] = "PLAY4TUNE — RTP todos os jogos desde D0"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Janela: {D0_OPERACAO.strftime('%d/%m/%Y')} a {DATA_FIM.strftime('%d/%m/%Y')} ({total_dias} dias)"
    ws["A3"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')} BRT"
    ws["A4"] = (f"Fonte: casino_user_game_metrics | "
                f"Moedas: BRL+PKR (1 PKR = R$ {mainmod.PKR_TO_BRL:.6f})")

    linhas = [
        ("Jogos com apostas no periodo", len(registros)),
        ("Jogos com amostra relevante (>=500 giros)", sum(1 for r in registros if r["rodadas"] >= 500)),
        ("ANOMALOS (delta > +15pp)",    sum(1 for r in registros if r["veredito"].startswith("ANOMALO"))),
        ("ALTOS (delta +5 a +15pp)",    sum(1 for r in registros if r["veredito"].startswith("alto"))),
        ("NORMAIS",                     sum(1 for r in registros if r["veredito"] == "normal")),
        ("ABAIXO (casa ganha)",         sum(1 for r in registros if r["veredito"].startswith("abaixo"))),
        ("Amostra baixa (<500 giros)",  sum(1 for r in registros if r["rodadas"] < 500)),
        ("Turnover total (BRL)",        f"R$ {to_brl(turnover_total):,.2f}"),
        ("Turnover total (PKR)",        f"Rs {turnover_total:,.2f}"),
        ("GGR casa (BRL)",              f"R$ {to_brl(ggr_total):,.2f}"),
        ("GGR casa (PKR)",              f"Rs {ggr_total:,.2f}"),
    ]
    for i, (k, v) in enumerate(linhas, start=6):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 28

    # Aba 2: Ranking completo (ordenado por delta desc)
    ws2 = wb.create_sheet("Ranking RTP")
    headers = [
        "Rank", "Jogo", "Provider", "Ativo",
        "RTP oficial 2J (%)", "RTP cfg banco (%)", "Delta banco vs oficial (pp)",
        "RTP obs (%)", "Delta vs ref (pp)",
        "Dias ativo", "Players", "Rodadas",
        "Apostado (BRL)", "Apostado (PKR)",
        "Ganho (BRL)", "Ganho (PKR)",
        "GGR (BRL)", "GGR (PKR)",
        "Veredito",
    ]
    for j, h in enumerate(headers, 1):
        c = ws2.cell(row=1, column=j, value=h)
        c.font = header_font; c.fill = header_fill; c.border = border
        c.alignment = Alignment(horizontal="center")

    # Ordenar: anomalos primeiro, depois altos, abaixo, normais, amostra baixa
    ordenados = sorted(registros, key=lambda r: (
        0 if r["veredito"].startswith("ANOMALO") else
        1 if r["veredito"].startswith("alto") else
        2 if r["veredito"].startswith("abaixo") else
        3 if r["veredito"] == "normal" else 4,
        -(r["delta"] or -999),
    ))

    for i, r in enumerate(ordenados, start=2):
        fill = None
        if r["veredito"].startswith("ANOMALO"):   fill = anomalo_fill
        elif r["veredito"].startswith("alto"):    fill = alto_fill
        elif r["veredito"].startswith("abaixo"):  fill = abaixo_fill
        row = [
            i - 1, r["jogo"], r["provider"], "Sim" if r["ativo"] else "Nao",
            r.get("rtp_oficial"), r["rtp_cfg"], r.get("delta_banco_vs_oficial"),
            r["rtp_obs"], r["delta"],
            r["dias_ativo"], r["players"], r["rodadas"],
            to_brl(r["apostado"]), r["apostado"],
            to_brl(r["ganho"]),    r["ganho"],
            to_brl(r["ggr"]),      r["ggr"],
            r["veredito"],
        ]
        for j, v in enumerate(row, 1):
            c = ws2.cell(row=i, column=j, value=v); c.border = border
            if fill: c.fill = fill
            if j in (5, 6, 7, 8, 9): c.number_format = '0.0'
            if j in (13, 14, 15, 16, 17, 18): c.number_format = '#,##0.00'

    widths = [6, 32, 18, 8, 12, 12, 14, 10, 12, 10, 10, 10, 14, 14, 14, 14, 14, 14, 22]
    for j, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(j)].width = w
    ws2.freeze_panes = "A2"

    # Aba 3: Legenda
    ws3 = wb.create_sheet("Legenda")
    ws3["A1"] = "Legenda e Criterios"
    ws3["A1"].font = Font(bold=True, size=14)
    lines = [
        ("", ""),
        ("RTP oficial 2J",         "Return to Player informado pela 2J Games na planilha oficial (source of truth)."),
        ("RTP cfg banco",          "Valor gravado em supernova_bet.casino_games.rtp (snapshot — pode estar defasado vs 2J)."),
        ("Delta banco vs oficial", "rtp_cfg_banco - rtp_oficial_2j. Positivo/negativo alto = catalogo do banco nao esta sincronizado com o provider."),
        ("RTP obs",                "Ganho total / Apostado total * 100 no periodo."),
        ("Delta vs ref (pp)",      "RTP obs - RTP de referencia (usa oficial 2J se houver; senao, banco). Positivo = casa devolveu mais que o esperado."),
        ("", ""),
        ("Criterios de classificacao", ""),
        ("ANOMALO",       "Delta > +15pp com amostra >=500 giros. Revisar urgente."),
        ("alto",          "Delta +5 a +15pp. Monitorar — pode ser volatilidade ou tendencia."),
        ("abaixo",        "Delta < -5pp. Casa ganha mais que o config prometia — investigar se RTP cfg esta correto."),
        ("normal",        "Delta entre -5 e +5pp. Oscilacao natural."),
        ("amostra baixa", "< 500 giros no periodo inteiro — ignorar, volatilidade alta."),
        ("", ""),
        ("Por que 500 giros?", "Com RTP ~95% tipico, 500 giros dao desvio padrao ~5pp. Abaixo disso o ruido mata a conclusao."),
        ("", ""),
        ("Fonte apostas",   "casino_user_game_metrics (supernova_bet) — validada 09/04/2026"),
        ("Fonte RTP banco", "supernova_bet.public.casino_games.rtp"),
        ("Fonte RTP 2J",    "Google Sheets 2J Games (source of truth) — export CSV em reports/rtp_oficial_2j_games.csv"),
        ("Filtros",         "72 contas teste excluidas (heuristica + logica dev), 4 whitelist DP/SQ devolvidas"),
        ("Moedas",          f"BRL + PKR (triangulacao USD via currency_exchange_rates)"),
        ("Taxa",  f"1 PKR = R$ {mainmod.PKR_TO_BRL:.6f} "
                  f"(snapshot {mainmod.FX_FETCHED_AT.strftime('%d/%m/%Y %H:%M UTC') if mainmod.FX_FETCHED_AT else ''})"),
    ]
    for i, (k, v) in enumerate(lines, start=2):
        ws3.cell(row=i, column=1, value=k).font = Font(bold=True) if k and not v else Font()
        ws3.cell(row=i, column=2, value=v)
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 85

    wb.save(XLSX_PATH)


if __name__ == "__main__":
    run()
    print("\nCONCLUIDO")
