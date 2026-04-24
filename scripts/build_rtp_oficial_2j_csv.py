"""Extrai RTP oficial 2J Games de TODAS as abas do 'API Games List.xlsx'.

Abas: Slots, Multiplayer, Crypto, TableGame, Fish.
RTP vem em formato variado:
  - Slots: "95.82%" (string com %)
  - Outros: 0.9769 (decimal) — multiplica por 100

Gera: reports/rtp_oficial_2j_games.csv
"""
import os, sys
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = r"C:\Users\NITRO\Downloads\API Games List.xlsx"
OUT  = os.path.join(ROOT, "reports", "rtp_oficial_2j_games.csv")

ABAS = ["Slots", "Multiplayer", "Crypto", "TableGame", "Fish"]


def find_col(header_row, targets):
    """Acha indice da coluna cujo header contem algum dos alvos (case-insensitive)."""
    for idx, cell in enumerate(header_row):
        if cell is None: continue
        val = str(cell).lower()
        for t in targets:
            if t.lower() in val:
                return idx
    return None


def normalize_rtp(raw):
    """Aceita '95.82%', '0.9582', 0.9582, 95.82 → retorna 95.82 (pct)."""
    if raw is None: return None
    if isinstance(raw, str):
        s = raw.replace(",", ".").replace("%", "").strip()
        if not s: return None
        try: n = float(s)
        except ValueError: return None
    else:
        try: n = float(raw)
        except (TypeError, ValueError): return None
    # Se < 10 assume decimal (0.9582) → vira pct
    if n < 10:
        n *= 100
    return round(n, 2)


def extract_games():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    rows = []
    for aba in ABAS:
        if aba not in wb.sheetnames:
            print(f"[SKIP] Aba '{aba}' nao existe")
            continue
        ws = wb[aba]
        # Header — tipicamente linha 1
        header = [c.value for c in ws[1]]
        col_id   = find_col(header, ["Game ID", "游戏ID"]) or 0
        col_name = find_col(header, ["Game Name", "游戏名称（英）"])
        col_rtp  = find_col(header, ["RTP", "返奖率"])
        if col_name is None or col_rtp is None:
            print(f"[WARN] {aba}: nao achou Game Name ({col_name}) ou RTP ({col_rtp}). Headers:")
            for i, h in enumerate(header[:12]):
                print(f"   [{i}] {h}")
            continue

        count = 0
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r[col_id] is None and r[col_name] is None:
                continue
            game_id = r[col_id]
            raw_name = r[col_name]
            if game_id is None or raw_name is None:
                continue
            # Game ID pode vir como float (ex: 1005.0) → normaliza p/ int string
            if isinstance(game_id, float) and game_id.is_integer():
                eid = str(int(game_id))
            else:
                eid = str(game_id).strip()
            # Nome: ja vem em ingles puro (a aba Slots tem em celula separada — col 2)
            name = " ".join(str(raw_name).split()).upper()  # colapsa quebras de linha / espacos
            rtp = normalize_rtp(r[col_rtp])
            if rtp is None:
                continue
            rows.append({
                "external_id": eid,
                "name": name,
                "rtp": rtp,
                "source": f"2J_API_Games_List_{aba}_20260424",
            })
            count += 1
        print(f"[{aba}] {count} jogos extraidos")
    return rows


def write_csv(rows):
    import csv
    # Dedup por external_id (mantem primeira aparicao)
    seen = set()
    deduped = []
    for row in rows:
        eid = row["external_id"]
        if eid in seen: continue
        seen.add(eid)
        deduped.append(row)
    with open(OUT, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["external_id", "name", "rtp", "source"])
        w.writeheader()
        for row in deduped:
            w.writerow(row)
    print(f"\n[CSV] {len(deduped)} jogos unicos salvos em: {OUT}")
    return deduped


if __name__ == "__main__":
    rows = extract_games()
    deduped = write_csv(rows)
    print(f"\nTotal por aba (pre-dedup): {len(rows)}")
    print(f"Total apos dedup por external_id: {len(deduped)}")
