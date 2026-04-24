"""
Report: Jogadores que estao afundando o GGR da Play4Tune — ultimos 7 dias

Demanda (20/04/2026): Head pediu para identificar quais jogadores estao
derrubando o GGR da P4T e quais jogos jogaram.

Janela: 13/04 a 19/04 (ultimos 7 dias COMPLETOS, D-1)
Motivo: 20/04 (hoje) e D-0 parcial — regra do time.

Fonte confiavel: casino_user_game_metrics (nao usar tabela bets — GGR errado)
Filtro: users reais (role=USER, sem padroes de teste)

Moeda original: PKR (Rupia Paquistanesa)
Moeda exibida: BRL — convertido via triangulacao USD usando a tabela
`currency_exchange_rates` do proprio supernova_bet (openexchangerate-free,
atualizada 2x/dia).

Entrega:
  - reports/play4_players_afundando_ggr_20abr_FINAL.xlsx (com aba Legenda)
  - reports/play4_players_afundando_ggr_20abr_resumo_wpp.txt
"""

import os
import sys
from datetime import datetime, date

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from db.supernova_bet import get_supernova_bet_connection

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ============================================================
# CONFIG
# ============================================================

DATA_INICIO = date(2026, 4, 13)   # segunda
DATA_FIM    = date(2026, 4, 19)   # domingo (D-1 em relacao a 20/04)
TOP_N       = 20                   # ranking principal

OUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "reports")
os.makedirs(OUT_DIR, exist_ok=True)

TS = datetime.now().strftime("%Y%m%d_%H%M")
XLSX_PATH = os.path.join(OUT_DIR, f"play4_players_afundando_ggr_20abr_FINAL.xlsx")
TXT_PATH  = os.path.join(OUT_DIR, f"play4_players_afundando_ggr_20abr_resumo_wpp.txt")


# Filtro de test users — mesma logica de listar_usuarios_teste_play4tune.py
# Whitelist DP/SQ — usuarios reais que a logica dev marcaria como teste
# (caso Deposito/Saque operacional ajustado manualmente pelo suporte)
# Fonte: memory/project_dp_sq_whitelist_p4t.md
REAL_USERS_WHITELIST = {
    'maharshani44377634693',
    'muhammadrehan17657797557',
    'rehmanzafar006972281',
    'saimkyani15688267',
}


# ============================================================
# FX — taxa PKR -> BRL via triangulacao USD (banco supernova_bet)
# ============================================================
# Var module-level, populada em run() apos conectar
PKR_TO_BRL = None
FX_FETCHED_AT = None


def fetch_fx_rates(cur):
    """Retorna (pkr_to_brl, fetched_at_utc). Triangulacao: BRL_per_PKR = BRL_per_USD / PKR_per_USD."""
    cur.execute("""
        WITH latest AS (
          SELECT DISTINCT ON (r.currency_id) c.code, r.rate, r.fetched_at
          FROM currency_exchange_rates r
          JOIN currencies c ON c.id = r.currency_id
          WHERE c.code IN ('BRL','PKR')
          ORDER BY r.currency_id, r.fetched_at DESC
        )
        SELECT
            (SELECT rate FROM latest WHERE code='BRL') AS brl_per_usd,
            (SELECT rate FROM latest WHERE code='PKR') AS pkr_per_usd,
            (SELECT MAX(fetched_at) FROM latest) AS fetched_at
    """)
    brl_per_usd, pkr_per_usd, fetched_at = cur.fetchone()
    pkr_to_brl = float(brl_per_usd) / float(pkr_per_usd)
    return pkr_to_brl, fetched_at


def to_brl(pkr_value):
    """Converte valor PKR para BRL usando a taxa cacheada (PKR_TO_BRL)."""
    if pkr_value is None:
        return 0.0
    return float(pkr_value) * PKR_TO_BRL


def get_test_user_ids(cur):
    """Retorna tuple de UUIDs de contas teste (heuristica UNION logica dev),
    excluindo whitelist DP/SQ. Lógica oficial: feedback_test_users_logica_dev_p4t.md"""
    cur.execute("""
        SELECT u.id, u.username
        FROM users u
        WHERE
           -- (a) HEURISTICA: padroes de username/email/role
           u.role != 'USER'
           OR LOWER(u.username) LIKE '%%test%%'
           OR LOWER(u.username) LIKE '%%teste%%'
           OR LOWER(u.username) LIKE '%%demo%%'
           OR LOWER(u.username) LIKE '%%admin%%'
           OR LOWER(COALESCE(u.email, '')) LIKE '%%@karinzitta%%'
           OR LOWER(COALESCE(u.email, '')) LIKE '%%@multi.bet%%'
           OR LOWER(COALESCE(u.email, '')) LIKE '%%@grupo-pgs%%'
           OR LOWER(COALESCE(u.email, '')) LIKE '%%@supernovagaming%%'
           OR LOWER(COALESCE(u.email, '')) LIKE '%%@play4tune%%'
           -- (b) LOGICA OFICIAL DEV: manipulacao manual de saldo ou deposito confirmado manual
           OR u.id IN (
               SELECT DISTINCT t.user_id
               FROM transactions t
               WHERE t.type IN ('ADJUSTMENT_CREDIT', 'ADJUSTMENT_DEBIT')
                  OR (t.type = 'DEPOSIT' AND t.reviewed_by IS NOT NULL)
           )
    """)
    rows = cur.fetchall()
    filtered = [r for r in rows if r[1] not in REAL_USERS_WHITELIST]
    whitelisted = [r for r in rows if r[1] in REAL_USERS_WHITELIST]
    ids = tuple(r[0] for r in filtered) or ('00000000-0000-0000-0000-000000000000',)
    return ids, len(filtered), whitelisted


# ============================================================
# QUERIES
# ============================================================

SQL_SANITY = """
SELECT
    MIN(m.date) AS primeiro_dia,
    MAX(m.date) AS ultimo_dia,
    COUNT(DISTINCT m.date) AS dias,
    COUNT(DISTINCT m.user_id) AS jogadores_ativos,
    ROUND(SUM(m.total_bet_amount)::numeric, 2) AS turnover_total,
    ROUND(SUM(m.total_win_amount)::numeric, 2) AS ganho_total,
    ROUND(SUM(m.net_revenue)::numeric, 2) AS ggr_total_casa
FROM casino_user_game_metrics m
WHERE m.date BETWEEN %s AND %s
  AND m.user_id NOT IN %s
"""


SQL_TOP_PLAYERS = """
SELECT
    u.username,
    u.public_id,
    u.phone,
    u.created_at::date AS cadastro,
    (%s::date - u.created_at::date) AS dias_conta,
    ROUND(SUM(m.total_bet_amount)::numeric, 2)  AS apostado_pkr,
    ROUND(SUM(m.total_win_amount)::numeric, 2)  AS ganho_pkr,
    ROUND(SUM(m.net_revenue)::numeric, 2)       AS ggr_pkr,
    SUM(m.played_rounds)                        AS rodadas,
    COUNT(DISTINCT m.game_id)                   AS jogos_unicos,
    COUNT(DISTINCT m.date)                      AS dias_jogou,
    ROUND(
        CASE WHEN SUM(m.total_bet_amount) > 0
             THEN SUM(m.total_win_amount) / SUM(m.total_bet_amount) * 100
             ELSE 0 END::numeric, 1
    ) AS payout_pct
FROM casino_user_game_metrics m
JOIN users u ON u.id = m.user_id
WHERE m.date BETWEEN %s AND %s
  AND m.user_id NOT IN %s
GROUP BY u.username, u.public_id, u.phone, u.created_at
HAVING SUM(m.net_revenue) < 0
ORDER BY SUM(m.net_revenue) ASC
LIMIT %s
"""


SQL_JOGOS_DO_PLAYER = f"""
SELECT
    g.name AS jogo,
    COALESCE(pv.name, 'Unknown') AS provider,
    g.rtp AS rtp_config,
    SUM(m.played_rounds) AS rodadas,
    ROUND(SUM(m.total_bet_amount)::numeric, 2)  AS apostado_pkr,
    ROUND(SUM(m.total_win_amount)::numeric, 2)  AS ganho_pkr,
    ROUND(SUM(m.net_revenue)::numeric, 2)       AS ggr_pkr,
    ROUND(
        CASE WHEN SUM(m.total_bet_amount) > 0
             THEN SUM(m.total_win_amount) / SUM(m.total_bet_amount) * 100
             ELSE 0 END::numeric, 1
    ) AS payout_pct
FROM casino_user_game_metrics m
JOIN users u            ON u.id = m.user_id
JOIN casino_games g     ON g.id = m.game_id
LEFT JOIN casino_providers pv ON pv.id = g.provider_id
WHERE m.date BETWEEN %s AND %s
  AND u.username = %s
GROUP BY g.name, pv.name, g.rtp
ORDER BY SUM(m.net_revenue) ASC
"""


# ============================================================
# FORMATACAO
# ============================================================

def fmt_pkr(v):
    if v is None:
        return "Rs 0,00"
    return f"Rs {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def fmt_brl(v_pkr):
    """Formata valor PKR como BRL usando a taxa cacheada."""
    if v_pkr is None:
        return "R$ 0,00"
    brl = to_brl(v_pkr)
    return f"R$ {brl:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def fmt_brl_pkr(v_pkr):
    """Formato dual: 'R$ X,XX (Rs Y,YY)' para exibir BRL + PKR lado a lado."""
    return f"{fmt_brl(v_pkr)} ({fmt_pkr(v_pkr)})"


def fmt_int(v):
    if v is None:
        return "0"
    return f"{int(v):,}".replace(",", ".")


# ============================================================
# MAIN
# ============================================================

def run():
    print("=" * 80)
    print("PLAY4TUNE — Jogadores que estao afundando o GGR")
    print(f"Janela: {DATA_INICIO.strftime('%d/%m/%Y')} a {DATA_FIM.strftime('%d/%m/%Y')} (7 dias)")
    print(f"Relatorio gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print("=" * 80)

    tunnel, conn = get_supernova_bet_connection()
    conn.set_session(readonly=True, autocommit=True)
    cur = conn.cursor()

    try:
        # ---------- Taxa PKR -> BRL (via triangulacao USD no banco) ----------
        global PKR_TO_BRL, FX_FETCHED_AT
        PKR_TO_BRL, FX_FETCHED_AT = fetch_fx_rates(cur)
        print(f"[FX] 1 PKR = R$ {PKR_TO_BRL:.6f} (fonte: openexchangerate-free @ {FX_FETCHED_AT} UTC)")

        # ---------- Identificar test users ----------
        test_ids, n_test, whitelisted = get_test_user_ids(cur)
        print(f"[Filtro] {n_test} contas teste excluidas (heuristica UNION logica dev)")
        if whitelisted:
            print(f"[Whitelist DP/SQ] {len(whitelisted)} contas reais devolvidas: "
                  f"{', '.join(r[1] for r in whitelisted)}")

        # ---------- Sanity ----------
        cur.execute(SQL_SANITY, (DATA_INICIO, DATA_FIM, test_ids))
        s = cur.fetchone()
        primeiro, ultimo, dias, ativos, turnover, ganho, ggr_casa = s
        print(f"\n[Sanity] Periodo banco: {primeiro} a {ultimo} ({dias} dias com dado)")
        print(f"[Sanity] Jogadores reais ativos: {ativos}")
        print(f"[Sanity] Turnover total: {fmt_brl_pkr(turnover)}")
        print(f"[Sanity] Ganho jogadores: {fmt_brl_pkr(ganho)}")
        print(f"[Sanity] GGR casa (periodo): {fmt_brl_pkr(ggr_casa)}")

        # ---------- Top players ----------
        cur.execute(SQL_TOP_PLAYERS, (DATA_FIM, DATA_INICIO, DATA_FIM, test_ids, TOP_N))
        top_players = cur.fetchall()
        print(f"\n[Top {TOP_N}] Jogadores com GGR negativo (casa perdeu):")

        if not top_players:
            print("  Nenhum jogador com GGR negativo no periodo.")
            return

        players_detail = []
        for i, p in enumerate(top_players, 1):
            username, pid, phone, cadastro, dias_conta, apostado, ganho, ggr, rodadas, jogos_u, dias_j, payout = p
            print(f"  {i:>2}. {username:<20} | PID {pid} | {dias_conta}d | "
                  f"Apostou {fmt_brl_pkr(apostado)} | GGR {fmt_brl_pkr(ggr)} | "
                  f"Payout {payout}%")

            # Buscar jogos do player
            cur.execute(SQL_JOGOS_DO_PLAYER, (DATA_INICIO, DATA_FIM, username))
            jogos = cur.fetchall()
            players_detail.append({
                "rank": i,
                "username": username,
                "public_id": pid,
                "phone": phone,
                "cadastro": cadastro,
                "dias_conta": dias_conta,
                "apostado_pkr": float(apostado or 0),
                "ganho_pkr": float(ganho or 0),
                "ggr_pkr": float(ggr or 0),
                "rodadas": int(rodadas or 0),
                "jogos_unicos": int(jogos_u or 0),
                "dias_jogou": int(dias_j or 0),
                "payout_pct": float(payout or 0),
                "jogos": jogos,
            })

        # ---------- Excel ----------
        if HAS_OPENPYXL:
            gerar_excel(players_detail, s)
            print(f"\n[Excel] Salvo em: {XLSX_PATH}")
        else:
            print("[Excel] openpyxl nao instalado, pulando.")

        # ---------- WhatsApp ----------
        gerar_resumo_whatsapp(players_detail, s)
        print(f"[WhatsApp] Resumo em: {TXT_PATH}")

    finally:
        cur.close()
        conn.close()
        tunnel.stop()

    print("\n" + "=" * 80)
    print("CONCLUIDO")
    print("=" * 80)


# ============================================================
# EXCEL
# ============================================================

def gerar_excel(players_detail, sanity):
    wb = openpyxl.Workbook()

    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    critico_fill = PatternFill("solid", fgColor="FEE2E2")
    alerta_fill  = PatternFill("solid", fgColor="FEF3C7")

    # ---------- Aba 1: Resumo ----------
    ws = wb.active
    ws.title = "Resumo"
    ws["A1"] = "PLAY4TUNE — Jogadores afundando o GGR"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Periodo: {DATA_INICIO.strftime('%d/%m/%Y')} a {DATA_FIM.strftime('%d/%m/%Y')}"
    ws["A3"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')} BRT"
    ws["A4"] = (f"Fonte: casino_user_game_metrics (supernova_bet) | "
                f"Moeda exibida: BRL (convertido de PKR @ 1 PKR = R$ {PKR_TO_BRL:.6f}, "
                f"taxa do banco {FX_FETCHED_AT.strftime('%d/%m/%Y %H:%M') if FX_FETCHED_AT else ''} UTC)")

    ws["A6"] = "Metrica"
    ws["B6"] = "Valor"
    ws["A6"].font = header_font
    ws["B6"].font = header_font
    ws["A6"].fill = header_fill
    ws["B6"].fill = header_fill

    resumo = [
        ("Jogadores reais ativos", fmt_int(sanity[3])),
        ("Turnover total (apostado)", fmt_brl_pkr(sanity[4])),
        ("Ganho total dos jogadores", fmt_brl_pkr(sanity[5])),
        ("GGR da casa", fmt_brl_pkr(sanity[6])),
        ("Jogadores com GGR negativo (top N exibido)", str(len(players_detail))),
        ("GGR dos top exibidos (impacto na casa)",
         fmt_brl_pkr(sum(p["ggr_pkr"] for p in players_detail))),
        ("Taxa PKR -> BRL usada", f"1 PKR = R$ {PKR_TO_BRL:.6f}"),
    ]
    for i, (k, v) in enumerate(resumo, start=7):
        ws[f"A{i}"] = k
        ws[f"B{i}"] = v
        ws[f"A{i}"].border = border
        ws[f"B{i}"].border = border

    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 28

    # ---------- Aba 2: Top Players ----------
    ws2 = wb.create_sheet("Top Players")
    headers = [
        "Rank", "Username", "Public ID", "Phone", "Cadastro", "Dias conta",
        "Apostado (BRL)", "Apostado (PKR)",
        "Ganho (BRL)",    "Ganho (PKR)",
        "GGR casa (BRL)", "GGR casa (PKR)",
        "Rodadas", "Jogos unicos", "Dias jogou", "Payout %",
        "Flag",
    ]
    for j, h in enumerate(headers, 1):
        c = ws2.cell(row=1, column=j, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = border

    for i, p in enumerate(players_detail, start=2):
        ggr_brl = to_brl(p["ggr_pkr"])
        # Faixas em BRL (aprox: 10K PKR ~ R$ 180, 5K PKR ~ R$ 90)
        if ggr_brl < -180:
            flag, fill = "CRITICO", critico_fill
        elif ggr_brl < -90:
            flag, fill = "ALERTA", alerta_fill
        else:
            flag, fill = "ATENCAO", None

        row_values = [
            p["rank"], p["username"], p["public_id"], p["phone"],
            p["cadastro"].strftime("%d/%m/%Y") if p["cadastro"] else "",
            p["dias_conta"],
            to_brl(p["apostado_pkr"]), float(p["apostado_pkr"]),
            to_brl(p["ganho_pkr"]),    float(p["ganho_pkr"]),
            ggr_brl,                   float(p["ggr_pkr"]),
            p["rodadas"], p["jogos_unicos"], p["dias_jogou"], p["payout_pct"],
            flag,
        ]
        for j, v in enumerate(row_values, 1):
            c = ws2.cell(row=i, column=j, value=v)
            c.border = border
            if fill:
                c.fill = fill
            if j in (7, 8, 9, 10, 11, 12):  # colunas monetarias BRL/PKR
                c.number_format = '#,##0.00'
            if j == 16:  # Payout %
                c.number_format = '0.0'

    # Larguras
    widths = [6, 22, 12, 18, 12, 8, 14, 14, 14, 14, 14, 14, 10, 10, 10, 10, 10]
    for j, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(j)].width = w

    # ---------- Aba 3: Detalhe por jogo ----------
    ws3 = wb.create_sheet("Detalhe por Jogo")
    det_headers = [
        "Rank", "Username", "Jogo", "Provider", "RTP config",
        "Rodadas",
        "Apostado (BRL)", "Apostado (PKR)",
        "Ganho (BRL)",    "Ganho (PKR)",
        "GGR casa (BRL)", "GGR casa (PKR)",
        "Payout %",
    ]
    for j, h in enumerate(det_headers, 1):
        c = ws3.cell(row=1, column=j, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = border

    row_i = 2
    for p in players_detail:
        for jogo in p["jogos"]:
            name, provider, rtp_cfg, rodadas, apostado, ganho, ggr, payout = jogo
            row_values = [
                p["rank"], p["username"], name, provider,
                float(rtp_cfg) if rtp_cfg is not None else 0,
                int(rodadas or 0),
                to_brl(apostado), float(apostado or 0),
                to_brl(ganho),    float(ganho or 0),
                to_brl(ggr),      float(ggr or 0),
                float(payout or 0),
            ]
            for j, v in enumerate(row_values, 1):
                c = ws3.cell(row=row_i, column=j, value=v)
                c.border = border
                if j in (7, 8, 9, 10, 11, 12):  # monetarias BRL/PKR
                    c.number_format = '#,##0.00'
                if j in (5, 13):  # RTP, Payout
                    c.number_format = '0.0'
            row_i += 1

    widths3 = [6, 22, 28, 18, 10, 10, 14, 14, 14, 14, 14, 14, 10]
    for j, w in enumerate(widths3, 1):
        ws3.column_dimensions[get_column_letter(j)].width = w

    # ---------- Aba 4: Legenda ----------
    ws4 = wb.create_sheet("Legenda")
    ws4["A1"] = "Legenda e Dicionario"
    ws4["A1"].font = Font(bold=True, size=14)

    fx_stamp = FX_FETCHED_AT.strftime("%d/%m/%Y %H:%M UTC") if FX_FETCHED_AT else "—"
    legenda = [
        ("", ""),
        ("Como ler este relatorio", ""),
        ("", ""),
        ("O relatorio mostra jogadores REAIS (test users excluidos) que tiveram",  ""),
        ("GGR negativo para a casa no periodo analisado.",                         ""),
        ("GGR negativo = jogador ganhou mais do que apostou = casa perdeu dinheiro.",""),
        ("", ""),
        ("Glossario", ""),
        ("Turnover", "Valor total apostado pelo jogador (BRL)"),
        ("Ganho", "Valor total ganho pelo jogador nas rodadas (BRL)"),
        ("GGR casa", "Gross Gaming Revenue = Apostado - Ganho. Negativo = casa perdeu."),
        ("Rodadas", "Numero de giros/rounds jogados"),
        ("Payout %", "Ganho / Apostado * 100. RTP configurado costuma ser 95-97%."),
        ("Dias conta", "Dias desde o cadastro ate o ultimo dia do periodo"),
        ("Flag", "CRITICO (GGR < -R$ 180), ALERTA (< -R$ 90), ATENCAO (resto)"),
        ("", ""),
        ("Fonte dos dados", ""),
        ("Banco", "Super Nova Bet DB (supernova_bet) — PostgreSQL"),
        ("Tabela", "casino_user_game_metrics (validada como fonte confiavel em 09/04)"),
        ("Moeda original", "PKR (Rupia Paquistanesa) — operacao Play4Tune"),
        ("Moeda exibida", "BRL (Real brasileiro)"),
        ("Taxa usada", f"1 PKR = R$ {PKR_TO_BRL:.6f} (triangulacao USD via currency_exchange_rates)"),
        ("Fonte taxa", f"openexchangerate-free, snapshot {fx_stamp}"),
        ("", ""),
        ("Acao sugerida", ""),
        ("CRITICO/ALERTA", "Revisar com Riscos: jogador valido? padrao fraude? conta nova?"),
        ("Jogos com payout >100%", "Possivel vulnerabilidade no jogo/provider — revisar RTP"),
        ("Conta nova (<3d) + alto volume", "Avaliar flag fraude"),
    ]
    for i, (k, v) in enumerate(legenda, start=2):
        ws4.cell(row=i, column=1, value=k).font = Font(bold=True) if v == "" and k else Font()
        ws4.cell(row=i, column=2, value=v)

    ws4.column_dimensions["A"].width = 36
    ws4.column_dimensions["B"].width = 70

    wb.save(XLSX_PATH)


# ============================================================
# RESUMO WHATSAPP
# ============================================================

def gerar_resumo_whatsapp(players_detail, sanity):
    lines = []
    lines.append("*PLAY4TUNE — Top jogadores afundando o GGR*")
    lines.append(f"Periodo: {DATA_INICIO.strftime('%d/%m')} a {DATA_FIM.strftime('%d/%m/%Y')}")
    lines.append(f"Moedas: BRL + PKR (1 PKR = R$ {PKR_TO_BRL:.6f})")
    lines.append(f"Fonte: casino_user_game_metrics\n")

    lines.append(f"*GGR da casa no periodo:* {fmt_brl_pkr(sanity[6])}")
    lines.append(f"*Turnover total:* {fmt_brl_pkr(sanity[4])}")
    lines.append(f"*Jogadores ativos (reais):* {fmt_int(sanity[3])}")
    lines.append(f"*Com GGR negativo (exibidos):* {len(players_detail)}\n")

    lines.append("*TOP 10 — jogadores que mais afundaram:*")
    for p in players_detail[:10]:
        lines.append(
            f"{p['rank']:>2}. {p['username']} (PID {p['public_id']}, {p['dias_conta']}d conta)"
        )
        lines.append(
            f"    Apostou {fmt_brl_pkr(p['apostado_pkr'])}"
        )
        lines.append(
            f"    Ganhou  {fmt_brl_pkr(p['ganho_pkr'])}"
        )
        lines.append(
            f"    *GGR casa: {fmt_brl_pkr(p['ggr_pkr'])}* | {p['jogos_unicos']} jogos | payout {p['payout_pct']}%"
        )
        for jogo in p["jogos"][:2]:
            name, prov, rtp_cfg, rodadas, apostado, ganho, ggr, payout = jogo
            lines.append(
                f"      - {name} ({prov}): {fmt_int(rodadas)} giros, "
                f"GGR {fmt_brl_pkr(ggr)}, payout {payout}%"
            )
        lines.append("")

    with open(TXT_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


if __name__ == "__main__":
    run()
