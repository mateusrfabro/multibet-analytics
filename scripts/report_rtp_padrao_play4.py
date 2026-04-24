"""
Report: RTP observado por dia dos jogos que apareceram nos top 10 GGR negativo.

Demanda Head (20/04/2026): "puxa so o RTP desses jogos ai, por dia, ultimos 7 dias.
e ve se e um padrao."

Jogos analisados = uniao dos 2 top 10 (periodo 13-19/04 + D-0 20/04).
Janela: 13/04 a 19/04 (7 dias completos, D-1).

Metricas por jogo/dia:
  - Rodadas (amostra)
  - Apostado (turnover)
  - Ganho (pago aos jogadores)
  - RTP observado = ganho / apostado * 100
  - Desvio vs RTP configurado (catalogo)

Padrao detectado:
  - RTP observado consistentemente > configurado -> anomalia
  - Spikes em 1 dia -> provavelmente 1 jogador ganhando grande
  - Volume baixo + RTP alto -> amostra pequena, ignorar
"""

import os
import sys
from datetime import date

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from db.supernova_bet import get_supernova_bet_connection
import report_players_afundando_ggr_play4 as mainmod
from report_players_afundando_ggr_play4 import (
    get_test_user_ids, fetch_fx_rates, to_brl, fmt_int,
)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


DATA_INICIO = date(2026, 4, 13)
DATA_FIM    = date(2026, 4, 19)

# Uniao dos 2 top 10 (periodo 7d + D-0)
JOGOS_ALVO = sorted(set([
    # Top 10 periodo 7 dias (13-19/04)
    "ZEUS POWER", "JIN JI MAHJONG", "88 FORTUNES", "BINGO LUCKY CLOVER",
    "MYTHICAL ANIMALS", "777 FRUIT", "POWER OF KRAKEN", "GOLDEN CENTURY",
    "MONEY MANIA", "GOLDEN BUFFALO",
    # Top 10 periodo D-0 (20/04) — so pra ver historico nos 7 dias anteriores
    "VORTEX", "WILD BOUNTY SHOWDOWN", "STARLIGHT PRINCESS 1000", "SUPER ACE",
    "CLEOPATRA'S FLAME", "DRAGON TIGER", "MR.FRANKENSTEIN",
    "DOOM BULLET", "ELVES",
]))

OUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "reports")
XLSX_PATH = os.path.join(OUT_DIR, "play4_rtp_padrao_7dias_FINAL.xlsx")


SQL_RTP_POR_DIA = """
SELECT
    g.name AS jogo,
    g.rtp AS rtp_cfg,
    m.date AS dia,
    COUNT(DISTINCT m.user_id) AS players,
    SUM(m.played_rounds) AS rodadas,
    ROUND(SUM(m.total_bet_amount)::numeric, 2) AS apostado,
    ROUND(SUM(m.total_win_amount)::numeric, 2) AS ganho,
    ROUND(SUM(m.net_revenue)::numeric, 2) AS ggr,
    ROUND(CASE WHEN SUM(m.total_bet_amount) > 0
               THEN SUM(m.total_win_amount) / SUM(m.total_bet_amount) * 100
               ELSE NULL END::numeric, 1) AS rtp_obs
FROM casino_user_game_metrics m
JOIN casino_games g ON g.id = m.game_id
WHERE m.date BETWEEN %s AND %s
  AND m.user_id NOT IN %s
  AND g.name = ANY(%s)
GROUP BY g.name, g.rtp, m.date
ORDER BY g.name, m.date
"""


def classificar(rtp_obs, rtp_cfg, rodadas):
    """Retorna veredito pra linha."""
    if rtp_obs is None:
        return "sem dado"
    if rodadas is None or rodadas < 30:
        return "amostra baixa (<30)"
    delta = float(rtp_obs) - float(rtp_cfg or 0)
    if delta > 30:
        return "ANOMALO (+30 pp)"
    if delta > 15:
        return "alto (+15 pp)"
    if delta > 5:
        return "acima (+5 pp)"
    if delta < -5:
        return "baixo (casa ganhou)"
    return "normal"


def run():
    print("=" * 80)
    print("PLAY4TUNE — RTP observado por dia (ultimos 7 dias)")
    print(f"Janela: {DATA_INICIO} a {DATA_FIM}")
    print(f"Jogos alvo ({len(JOGOS_ALVO)}): {', '.join(JOGOS_ALVO)}")
    print("=" * 80)

    tunnel, conn = get_supernova_bet_connection()
    conn.set_session(readonly=True, autocommit=True)
    cur = conn.cursor()

    try:
        pkr_to_brl, fx_at = fetch_fx_rates(cur)
        mainmod.PKR_TO_BRL = pkr_to_brl
        mainmod.FX_FETCHED_AT = fx_at
        print(f"[FX] 1 PKR = R$ {pkr_to_brl:.6f} (snapshot {fx_at} UTC)\n")

        test_ids, _, _ = get_test_user_ids(cur)

        cur.execute(SQL_RTP_POR_DIA, (DATA_INICIO, DATA_FIM, test_ids, JOGOS_ALVO))
        rows = cur.fetchall()

        # Indexar por jogo
        por_jogo = {}
        for r in rows:
            jogo, rtp_cfg, dia, players, rodadas, apostado, ganho, ggr, rtp_obs = r
            por_jogo.setdefault(jogo, {"rtp_cfg": float(rtp_cfg or 0), "dias": []})
            por_jogo[jogo]["dias"].append({
                "dia": dia, "players": players, "rodadas": rodadas or 0,
                "apostado": float(apostado or 0), "ganho": float(ganho or 0),
                "ggr": float(ggr or 0), "rtp_obs": float(rtp_obs) if rtp_obs is not None else None,
            })

        # Stdout — resumo por jogo
        print(f"{'Jogo':<26} {'RTP cfg':>8} {'Dias':>4} {'Rodadas':>8} "
              f"{'Apost.(Rs)':>12} {'GGR (Rs)':>12} {'RTP obs med':>12} {'Veredito'}")
        print("-" * 110)

        sumarios = []
        for jogo in JOGOS_ALVO:
            d = por_jogo.get(jogo)
            if not d:
                print(f"{jogo:<26} {'—':>8} {'0':>4} — sem apostas no periodo")
                continue
            rtp_cfg = d["rtp_cfg"]
            dias = d["dias"]
            tot_rodadas = sum(x["rodadas"] for x in dias)
            tot_apostado = sum(x["apostado"] for x in dias)
            tot_ganho = sum(x["ganho"] for x in dias)
            tot_ggr = sum(x["ggr"] for x in dias)
            rtp_obs_med = (tot_ganho / tot_apostado * 100) if tot_apostado > 0 else None
            veredito = classificar(rtp_obs_med, rtp_cfg, tot_rodadas)

            print(f"{jogo:<26} {rtp_cfg:>8.2f} {len(dias):>4} {tot_rodadas:>8} "
                  f"{tot_apostado:>12,.0f} {tot_ggr:>12,.0f} "
                  f"{(rtp_obs_med if rtp_obs_med else 0):>11.1f}% {veredito}")

            sumarios.append({
                "jogo": jogo, "rtp_cfg": rtp_cfg, "dias_com_dado": len(dias),
                "rodadas": tot_rodadas, "apostado": tot_apostado, "ganho": tot_ganho,
                "ggr": tot_ggr, "rtp_obs_med": rtp_obs_med, "veredito": veredito,
                "dias_detalhe": dias,
            })

        # Detalhe por dia (so jogos anomalos / alerta)
        print("\n" + "=" * 80)
        print("DETALHE POR DIA — jogos com RTP obs medio ANOMALO ou ALTO (>+15pp)")
        print("=" * 80)
        for s in sumarios:
            if s["rtp_obs_med"] is None:
                continue
            delta = s["rtp_obs_med"] - s["rtp_cfg"]
            if delta <= 5:
                continue
            print(f"\n{s['jogo']} (RTP cfg {s['rtp_cfg']}%, obs medio "
                  f"{s['rtp_obs_med']:.1f}%, delta {delta:+.1f}pp):")
            for dd in s["dias_detalhe"]:
                print(f"  {dd['dia']} | {dd['players']:>2}p | {dd['rodadas']:>6} giros | "
                      f"Rs {dd['apostado']:>10,.0f} apost. | Rs {dd['ggr']:>10,.0f} GGR | "
                      f"RTP obs: "
                      f"{(dd['rtp_obs'] if dd['rtp_obs'] is not None else 0):>6.1f}%")

        # Excel
        if HAS_OPENPYXL:
            gerar_excel(sumarios, fx_at)
            print(f"\n[Excel] Salvo em: {XLSX_PATH}")

        return sumarios

    finally:
        cur.close(); conn.close(); tunnel.stop()


def gerar_excel(sumarios, fx_at):
    wb = openpyxl.Workbook()
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    anomalo_fill = PatternFill("solid", fgColor="FEE2E2")
    alto_fill    = PatternFill("solid", fgColor="FEF3C7")

    # Aba 1: Resumo
    ws = wb.active
    ws.title = "Resumo"
    ws["A1"] = "PLAY4TUNE — RTP observado vs configurado (ultimos 7 dias)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Janela: {DATA_INICIO.strftime('%d/%m/%Y')} a {DATA_FIM.strftime('%d/%m/%Y')}"
    ws["A3"] = f"Fonte: casino_user_game_metrics (supernova_bet) | Filtro: test users excluidos"

    headers = [
        "Jogo", "RTP config (%)", "Dias c/ dado", "Rodadas", "Apostado (PKR)",
        "Ganho (PKR)", "GGR (PKR)", "RTP obs medio (%)", "Delta (pp)", "Veredito",
    ]
    row0 = 5
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=row0, column=j, value=h)
        c.font = header_font; c.fill = header_fill; c.border = border
        c.alignment = Alignment(horizontal="center")

    for i, s in enumerate(sumarios, start=row0 + 1):
        delta = (s["rtp_obs_med"] - s["rtp_cfg"]) if s["rtp_obs_med"] is not None else None
        fill = None
        if s["veredito"].startswith("ANOMALO"):
            fill = anomalo_fill
        elif s["veredito"].startswith("alto"):
            fill = alto_fill
        row = [
            s["jogo"], s["rtp_cfg"], s["dias_com_dado"], s["rodadas"],
            s["apostado"], s["ganho"], s["ggr"],
            s["rtp_obs_med"], delta, s["veredito"],
        ]
        for j, v in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=v); c.border = border
            if fill: c.fill = fill
            if j in (2, 8, 9): c.number_format = '0.0'
            if j in (5, 6, 7): c.number_format = '#,##0.00'

    widths = [28, 14, 12, 12, 16, 16, 16, 16, 12, 22]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # Aba 2: Detalhe dia a dia (pivot jogo x dia)
    ws2 = wb.create_sheet("RTP por dia")
    dias_col = [DATA_INICIO.fromordinal(DATA_INICIO.toordinal() + i)
                for i in range((DATA_FIM - DATA_INICIO).days + 1)]
    headers2 = ["Jogo", "RTP config"] + [d.strftime("%d/%m") for d in dias_col] + ["RTP obs medio"]
    for j, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=j, value=h)
        c.font = header_font; c.fill = header_fill; c.border = border
        c.alignment = Alignment(horizontal="center")

    for i, s in enumerate(sumarios, start=2):
        rtp_by_day = {dd["dia"]: dd["rtp_obs"] for dd in s["dias_detalhe"]}
        row = [s["jogo"], s["rtp_cfg"]]
        for d in dias_col:
            val = rtp_by_day.get(d)
            row.append(val if val is not None else "")
        row.append(s["rtp_obs_med"])
        for j, v in enumerate(row, 1):
            c = ws2.cell(row=i, column=j, value=v); c.border = border
            if j >= 2 and isinstance(v, (int, float)):
                c.number_format = '0.0'
                # Color scale manual: vermelho se > 120, amarelo se > 105
                if isinstance(v, float) and v > 120:
                    c.fill = anomalo_fill
                elif isinstance(v, float) and v > 105:
                    c.fill = alto_fill

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 12
    for i in range(3, 3 + len(dias_col) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 10

    # Aba 3: Legenda
    ws3 = wb.create_sheet("Legenda")
    ws3["A1"] = "Como ler"
    ws3["A1"].font = Font(bold=True, size=14)
    lines = [
        ("RTP config", "Return to Player configurado no catalogo do provider (2J Games)"),
        ("RTP obs medio", "Ganho total / Apostado total * 100 (no periodo de 7 dias)"),
        ("Delta (pp)", "RTP obs - RTP config. Positivo = casa devolveu mais do que deveria."),
        ("Veredito ANOMALO", "Delta > +30 pp — provavel bug/fraude/config errada"),
        ("Veredito alto",    "Delta entre +15 e +30 pp — investigar"),
        ("Veredito acima",   "Delta entre +5 e +15 pp — monitorar"),
        ("Veredito normal",  "Delta entre -5 e +5 pp — oscilacao natural"),
        ("amostra baixa",    "< 30 rodadas no periodo — ignorar, amostra pequena"),
        ("", ""),
        ("Fonte", "casino_user_game_metrics (supernova_bet)"),
        ("Filtros", "72 contas teste excluidas (heuristica + logica dev), 4 whitelist DP/SQ"),
    ]
    for i, (k, v) in enumerate(lines, start=2):
        ws3.cell(row=i, column=1, value=k).font = Font(bold=True) if k else Font()
        ws3.cell(row=i, column=2, value=v)
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 80

    wb.save(XLSX_PATH)


if __name__ == "__main__":
    run()
    print("\nCONCLUIDO")
